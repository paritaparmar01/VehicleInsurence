from .models import Users, CombinedInfo, PolicyIssue, InsuranceEnquiry, VehicleInformation, LoanEnquiry, Document
from django.http import HttpResponse
from django.shortcuts import render, redirect
from .models import *
from django.http import HttpResponse, HttpResponseRedirect
from datetime import datetime
import pandas as pd
#from django_pandas.io import read_frame
from openpyxl import Workbook
from .models import PolicyIssue
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook
from decimal import Decimal, DecimalException
from django.contrib.auth.decorators import login_required
# from .forms import *

def index(request):
    return render(request, 'base.html')

def home(request):
    return render(request, 'index.html')

def try_policy(request):
    return render(request, 'trypolicy.html')

def try_loan(request):
    return render(request, 'tryloan.html')
 
def search(request):
    data = InsuranceEnquiry.objects.all()
    data2 = VehicleInformation.objects.all()
    if request.method == "POST":
        vi = request.POST.get('vi')
        rec = VehicleInformation.objects.filter(vehicle_number = vi)
        return render(request, 'try.html', {'data':data, 'data2' : data2, "rec" : rec})
    return render(request, 'try.html',{'data':data, 'data2' : data2})
                                    
def try_button(request):
    data = InsuranceEnquiry.objects.all()
    data2 = VehicleInformation.objects.all()
    if request.method == 'POST':
        # Extract common data
        name = request.POST.get('name')
        number = request.POST.get('number')
        email = request.POST.get('email')

        # Save Insurance Enquiry
        enquiry = InsuranceEnquiry.objects.create(
            name=name,
            mobile=number,
            email=email
        )

        # Extract and save vehicle-related data
        mobile = number
        vehicle_numbers = request.POST.getlist('vehicle_number[]')
        rc_books = request.POST.getlist('rc_book_radio')
        previous_policies = request.POST.getlist('previous_policy_radio')
        end_dates = request.POST.getlist('end_date[]')

        # Extract and save RC Book images and Previous Policy images separately
        rc_book_images = request.FILES.getlist('rc_book_image[]')
        previous_policy_images = request.FILES.getlist(
            'previous_policy_image[]')

        # List to store VehicleInformation objects
        vehicle_list = []

        for i in range(len(vehicle_numbers)):
            vehicle_number = vehicle_numbers[i]
            rc_book = rc_books[i]
            end_date = end_dates[i]

            # Check if the index is within bounds
            if i < len(rc_book_images) and i < len(previous_policies) and i < len(previous_policy_images):
                rc_book_image = rc_book_images[i]
                previous_policy = previous_policies[i]
                previous_policy_image = previous_policy_images[i]

                vehicle = VehicleInformation(
                    # insurance_enquiry=enquiry,
                    mobile=number,
                    name=name,
                    email=email,
                    vehicle_number=vehicle_number,
                    rc_book=rc_book,
                    rc_book_image=rc_book_image,
                    previous_policy=previous_policy,
                    previous_policy_image=previous_policy_image,
                    end_date=end_date
                )

                vehicle_list.append(vehicle)

        # Use bulk_create to insert all records at once
        if VehicleInformation.objects.bulk_create(vehicle_list):
            # Redirect to the desired URL after successful form submission
            context = {"check": True}
            return render(request, 'try.html', context)

    return render(request, 'try.html',{'data':data, 'data2' : data2})
    

def register(request):
    if request.method == 'POST':
        # Extract common data
        email = request.POST.get('email')
        name = request.POST.get('name')
        password = request.POST.get('password')

        # Save Insurance Enquiry
        Users.objects.create(
            email=email,
            name = name,
            password = password
        )

        # context = {"check": True}

        return render(request, 'login.html')
    
    return render(request, 'register.html')


# def login(request):
#     if request.method == 'POST':
#         # Extract common data
#         email = request.POST.get('email')
#         password = request.POST.get('password')

#         # user = authenticate(request, username = username, password = password)

#         # if user is not None:
#         #     login(request, user)
#         try:
#             user = Users.objects.get(email = email, password = password)
#             if Users:
#                 request.session['Users']=user
#             return render(request, 'base.html')
#         except:
#             return render(request, 'login.html')
    
#     return render(request, 'login.html')

def login(request):
    if request.method == 'POST':
        # Extract common data
        email = request.POST.get('email')
        password = request.POST.get('password')

        # user = authenticate(request, username = username, password = password)

        # if user is not None:
        #     login(request, user)
        try:
            user = Users.objects.get(email = email, password = password)
            if Users:
                request.session['Users']=user
                return render(request, 'index.html')
        except:
                return render(request, 'login.html')
    
    return render(request, 'login.html')


# def login(request):
#     if request.method == 'POST':
#         # Extract common data
#         mobile = request.POST.get('mobile')

#         try:
#             rec = InsuranceEnquiry.objects.get(mobile=mobile)
#             if rec:
#                 request.session['isLoggedin'] = True
#                 return render(request, 'try.html', {'rec' : rec})
#         except:
#             return render(request, 'try.html')
    
#     return render(request, 'try.html')


def forgotPwd(request):
    if request.method == 'POST':
        # Extract common data
        email = request.POST.get('email')
        password = request.POST.get('password')

        # user = authenticate(request, username = username, password = password)

        # if user is not None:
        #     login(request, user)
        try:
            user = Users.objects.get(email=email)
            user.password = password
            user.save()
            
            return render(request, 'base.html')
        except:
            return render(request, 'forgotpwd.html')

    return render(request, 'forgotpwd.html')

# def insurance_enquiry_view(request):
#     # template_name = 'blog/enquiry.html'

#     if request.method == 'POST':
#         enquiry_form = InsuranceEnquiryForm(request.POST)
#         vehicle_form = VehicleInformationForm(request.POST, request.FILES)

#         if enquiry_form.is_valid() and vehicle_form.is_valid():
#             enquiry = enquiry_form.save()
#             vehicle = vehicle_form.save(commit=False)
#             vehicle.insurance_enquiry = enquiry
#             vehicle.save()
#             return render(request, 'insurance/success.html')

#     else:
#         enquiry_form = InsuranceEnquiryForm()
#         vehicle_form = VehicleInformationForm()

#     return render(request, template_name, {'enquiry_form': enquiry_form, 'vehicle_form': vehicle_form})


# def enquiry(request):
#     if request.method == 'POST':
#         # Extract common data
#         name = request.POST.get('name')
#         number = request.POST.get('number')
#         email = request.POST.get('email')

#         # Save Insurance Enquiry
#         enquiry = InsuranceEnquiry.objects.create(
#             name=name,
#             number=number,
#             email=email
#         )

#         # Extract vehicle-related data
#         vehicle_numbers = request.POST.getlist('vehicle_number[]')
#         rc_books = request.POST.getlist('rc_book[]')
#         previous_policies = request.POST.getlist('previous_policy[]')
#         end_dates = request.POST.getlist('end_date[]')

#         print("date + vc",end_dates, vehicle_numbers)

#         # Extract and save RC Book images and Previous Policy images separately
#         rc_book_images = request.FILES.getlist('rc_book_image[]')
#         previous_policy_images = request.FILES.getlist(
#             'previous_policy_image[]')

#         for i in range(len(vehicle_numbers)):
#             vehicle_number = vehicle_numbers[i]
#             rc_book = rc_books[i]
#             end_date = end_dates[i]
#             print("date + vc", end_date, vehicle_number)

#             # Check if the index is within bounds
#             if i < len(rc_book_images) and i < len(previous_policies) and i < len(previous_policy_images):
#                 rc_book_image = rc_book_images[i]
#                 previous_policy = previous_policies[i]
#                 previous_policy_image = previous_policy_images[i]

#                 VehicleInformation.objects.create(
#                     insurance_enquiry=enquiry,
#                     vehicle_number=vehicle_number,
#                     rc_book=rc_book,
#                     rc_book_image=rc_book_image,
#                     previous_policy=previous_policy,
#                     previous_policy_image=previous_policy_image,
#                     end_date=end_date
#                 )
#                 # Redirect to the desired URL after successful form submission
#             return render(request, 'base.html')

#         return render(request, 'enquiry.html')

def enquiry(request):
    data = InsuranceEnquiry.objects.all()
    if request.method == 'POST':
        number = request.POST.get('mobile')
        
        user = InsuranceEnquiry.objects.filter(mobile=number)
        
        if user.exists():
            user = InsuranceEnquiry.objects.get(mobile=number)
            if user:
                request.session['mobile'] = user.mobile
                request.session['name'] = user.name
            return HttpResponseRedirect('/enquiryoldcust/')
        else:
            return HttpResponseRedirect('/enquirynewcust/')
    return render(request, 'enquiry.html',{'data':data})





def enquiryOldCust(request):
    if request.method == 'POST':
        # Extract common data
        number = request.POST.get('number')

        user = InsuranceEnquiry.objects.get(mobile=number)
        if user is None:
            return HttpResponseRedirect('/enquiryoldcust/')

        # Extract and save vehicle-related data
        mobile = number
        vehicle_numbers = request.POST.getlist('vehicle_number[]')
        rc_books = request.POST.getlist('rc_book[]')
        previous_policies = request.POST.getlist('previous_policy[]')
        end_dates = request.POST.getlist('end_date[]')

        # Extract and save RC Book images and Previous Policy images separately
        rc_book_images = request.FILES.getlist('rc_book_image[]')
        previous_policy_images = request.FILES.getlist(
            'previous_policy_image[]')

        # List to store VehicleInformation objects
        vehicle_list = []

        for i in range(len(vehicle_numbers)):
            vehicle_number = vehicle_numbers[i]
            rc_book = rc_books[i]
            end_date = end_dates[i]

            # Check if the index is within bounds
            if i < len(rc_book_images) and i < len(previous_policies) and i < len(previous_policy_images):
                rc_book_image = rc_book_images[i]
                previous_policy = previous_policies[i]
                previous_policy_image = previous_policy_images[i]

                vehicle = VehicleInformation(
                    # insurance_enquiry=enquiry,
                    mobile=number,
                    vehicle_number=vehicle_number,
                    rc_book=rc_book,
                    rc_book_image=rc_book_image,
                    previous_policy=previous_policy,
                    previous_policy_image=previous_policy_image,
                    end_date=end_date
                )
            else:
                if rc_book_images[i] is not None:
                    rc_book_image = rc_book_images[i]
                else:
                    rc_book_image = "-"
                
                if previous_policies[i] is not None:
                    previous_policy = previous_policies[i]
                else:
                    previous_policy = "-"
                
                if previous_policy_images[i] is not None:
                    previous_policy_image = previous_policy_images[i]
                else:
                    previous_policy_image = "-"
                if end_date[i] is not None:
                     end_date= end_date[i]
                else:
                    end_date = "-"
                vehicle = VehicleInformation(
                    # insurance_enquiry=enquiry,
                    mobile=number,
                    vehicle_number=vehicle_number,
                    rc_book=rc_book,
                    rc_book_image=rc_book_image,
                    previous_policy=previous_policy,
                    previous_policy_image=previous_policy_image,
                    end_date=end_date
                )

                vehicle_list.append(vehicle)

        # Use bulk_create to insert all records at once
        if VehicleInformation.objects.bulk_create(vehicle_list):
            # Redirect to the desired URL after successful form submission
            context = {"check" : True}
            return render(request, 'enquiry_oldcust.html', context)

    return render(request, 'enquiry_oldcust.html')

def enquiryNewCust(request):
    data = InsuranceEnquiry.objects.all()
    data = VehicleInformation.objects.all()
    if request.method == 'POST':
        # Extract common data
        name = request.POST.get('name')
        number = request.POST.get('number')
        email = request.POST.get('email')

        # Save Insurance Enquiry
        enquiry = InsuranceEnquiry.objects.create(
            name=name,
            mobile=number,
            email=email
        )

        # Extract and save vehicle-related data
        mobile = number
        vehicle_numbers = request.POST.getlist('vehicle_number[]')
        rc_books = request.POST.getlist('rc_book_radio')
        previous_policies = request.POST.getlist('previous_policy_radio')
        end_dates = request.POST.getlist('end_date[]')

        # Extract and save RC Book images and Previous Policy images separately
        rc_book_images = request.FILES.getlist('rc_book_image[]')
        previous_policy_images = request.FILES.getlist(
            'previous_policy_image[]')

        # List to store VehicleInformation objects
        vehicle_list = []

        for i in range(len(vehicle_numbers)):
            vehicle_number = vehicle_numbers[i]
            rc_book = rc_books[i]
            end_date = end_dates[i]

            # Check if the index is within bounds
            if i < len(rc_book_images) and i < len(previous_policies) and i < len(previous_policy_images):
                rc_book_image = rc_book_images[i]
                previous_policy = previous_policies[i]
                previous_policy_image = previous_policy_images[i]

                vehicle = VehicleInformation(
                    # insurance_enquiry=enquiry,
                    mobile=number,
                    vehicle_number=vehicle_number,
                    rc_book=rc_book,
                    rc_book_image=rc_book_image,
                    previous_policy=previous_policy,
                    previous_policy_image=previous_policy_image,
                    end_date=end_date
                )

                vehicle_list.append(vehicle)

        # Use bulk_create to insert all records at once
        if VehicleInformation.objects.bulk_create(vehicle_list):
            # Redirect to the desired URL after successful form submission
            context = {"check": True}
            return render(request, 'enquiry_newcust.html', context)

    return render(request, 'enquiry_newcust.html',{'data':data})

def policy_issue(request):
    data = PolicyIssue.objects.all()
    if request.method == 'POST':
        date = request.POST.get('date')
        name = request.POST.get('name')
        number = request.POST.get('number')
        p_number = request.POST.get('p_number')
        v_number = request.POST.get('v_number')
        Vehicle = request.POST.get('Vehicle')
        c_number = request.POST.get('c_number')
        e_number = request.POST.get('e_number')
        Location = request.POST.get('Location')
        HP_bank = request.POST.get('HP_bank')
        business_type = request.POST.get('business_type', 'Data')
        insurance_type = request.POST.get('insurance_type', 'TP')
        insurance_portal = request.POST.get('insurance_portal', 'Agency')
        I_company = request.POST.get('I_company')
        payment = request.POST.get('payment')
        payment_sos = request.POST.get('payment_sos')
        PS_date = request.POST.get('PS_date')
        PE_date = request.POST.get('PE_date')
        Ncb = request.POST.get('Ncb')
        Premium = request.POST.get('Premium')
        odNetPremium = request.POST.get('odNetPremium')
        commissionPercentage = request.POST.get('commissionPercentage')
        profitResult = request.POST.get('profitResult')
        tdsPercentage = request.POST.get('tdsPercentage', 5)
        profitAfterTDSResult = request.POST.get('profitAfterTDSResult')
        PayoutAmount = request.POST.get('PayoutAmount')
        payoutDiscount = request.POST.get('payoutDiscount')
        netProfitResult = request.POST.get('netProfitResult')
        Executive = request.POST.get('Executive')
        DSA = request.POST.get('DSA')

        # Check if the date is provided and not empty
        if date:
            try:
                # Convert the date string to a datetime object
                date = datetime.strptime(date, '%Y-%m-%d').date()
            except ValueError:
                return HttpResponse('Invalid date format. Please use YYYY-MM-DD.')
        else:
            # return HttpResponse('Date is required.')
            date = datetime()
        
        

        PolicyIssue.objects.create(
            date=date,
            name=name,
            number=number,
            p_number=p_number,
            v_number=v_number,
            Vehicle=Vehicle,
            c_number=c_number,
            e_number=e_number,
            Location=Location,
            HP_bank=HP_bank,
            business_type=business_type,
            insurance_type=insurance_type,
            insurance_portal=insurance_portal,
            I_company=I_company,
            payment=payment,
            payment_sos=payment_sos,
            PS_date=PS_date,
            PE_date=PE_date,
            Ncb=Ncb,
            Premium=Premium,
            odNetPremium=odNetPremium,
            commissionPercentage=commissionPercentage,
            profitResult=profitResult,
            tdsPercentage=tdsPercentage,
            profitAfterTDSResult=profitAfterTDSResult,
            PayoutAmount=PayoutAmount,
            payoutDiscount=payoutDiscount,
            netProfitResult=netProfitResult,
            Executive=Executive,
            DSA=DSA
            
            
        )

        context = {"check" : True,}        # Redirect to the desired URL after successful form submission
        return render(request, 'policyissue.html', context=context)

    return render(request, 'policyissue.html',{'data':data})

def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Read the Excel file into a DataFrame using pandas
        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            return render(request, 'uploadexcel.html', {'error_message': f'Error reading Excel file: {e}'})

        # Loop through the DataFrame and create PolicyIssue objects
        for index, row in df.iterrows():
            try:
                PolicyIssue.objects.create(
                    date=row['date'],
                    name=row['name'],
                    number=row['number'],
                    p_number=row['p_number'],
                    v_number=row['v_number'],
                    Vehicle=row['Vehicle'],
                    c_number=row['c_number'],
                    e_number=row['e_number'],
                    Location=row['Location'],
                    HP_bank=row['HP_bank'],
                    business_type=row['business_type'],
                    insurance_type=row['insurance_type'],
                    insurance_portal=row['insurance_portal'],
                    I_company=row['I_company'],
                    payment=row['payment'],
                    payment_sos=row['payment_sos'],
                    PS_date=row['PS_date'],
                    PE_date=row['PE_date'],
                    Ncb=row['Ncb%'],
                    odNetPremium=row['odNetPremium'],
                    commissionPercentage=row['commissionPercentage'],
                    profit=row['profit'],
                    tdsPercentage=row['tdsPercentage'],
                    profit_after_tds=row['profit_after_tds'],
                    payout_discount=row['payout_discount'],
                    net_profit=row['net_profit'],
                    Executive=row['Executive'],
                    DSA=row['DSA']
                    # Include other fields here...
                )
            except Exception as e:
                return render(request, 'uploadexcel.html', {'error_message': f'Error creating PolicyIssue: {e}'})

        context = {"check": True}
        return render(request, 'uploadexcel.html', context=context)

    return render(request, 'uploadexcel.html')

def parse_date(date_str):
    formats = ['%Y-%m-%d', '%m-%d-%Y', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%Y']  # Add more formats as needed
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            pass
    raise ValueError('Invalid date format')

def loan(request):
    data = LoanEnquiry.objects.all()
    if request.method == 'POST':
        # Extract common data
        name = request.POST.get('name')
        number = request.POST.get('number')
        email = request.POST.get('email')

        # Save Loan Enquiry
        loan_enquiry = LoanEnquiry.objects.create(
            name=name,
            number=number,
            email=email
        )

        # Extract and save document-related data
        rc_books = request.POST.getlist('rc_book[]')
        rc_book_images = request.FILES.getlist('rc_book_image[]')
        documents = request.FILES.getlist('documents[]')

        for rc_book, rc_book_image, document in zip(rc_books, rc_book_images, documents):
            Document.objects.create(
                loan_enquiry=loan_enquiry,
                rc_book=rc_book,
                rc_book_image=rc_book_image,
                document=document
            )

        context = {"check": True}
        return render(request, 'loan.html', context=context)

    return render(request, 'loan.html',{'data':data})
import pandas as pd
from django.shortcuts import render
from .models import PolicyIssue

def upload_policy(request):
    if request.method == 'POST':
        xlsx_file = request.FILES.get('XLSPolicy')

        if not xlsx_file:
            return HttpResponse('No file uploaded')

        try:
            wb = load_workbook(xlsx_file, read_only=True)
            ws = wb.active
            
            # Assuming the first row contains headers
            headers = [cell.value for cell in ws[1]]
            
            # Mapping headers to column indices
            header_indices = {header: index for index, header in enumerate(headers)}
            
            # Required columns
            required_columns = ['date', 'name', 'number', 'p_number', 'v_number', 'Vehicle', 
                                'c_number', 'e_number', 'Location', 'HP_bank', 'business_type', 
                                'insurance_type', 'insurance_portal', 'I_company', 'payment', 
                                'payment_sos', 'PS_date', 'PE_date', 'Ncb', 'Premium', 
                                'odNetPremium', 'commissionPercentage', 'profitResult', 'tdsPercentage', 
                                'profitAfterTDSResult','payoutDiscount','PayoutAmount', 'netProfitResult', 
                                'Executive', 'DSA']
            
            for column in required_columns:
                if column not in header_indices:
                    return HttpResponse(f'Error: Column "{column}" not found in the Excel file')
            
            for row in ws.iter_rows(min_row=2):
                # Initialize a dictionary to store cell values
                row_data = {}
                for header, index in header_indices.items():
                    # Extract cell value
                    cell_value = row[index].value
                    # Check if cell value is a formula or expression
                    if isinstance(cell_value, str) and cell_value.startswith('='):
                        # If it's a formula, set it to None for now
                        cell_value = None
                    row_data[header] = cell_value
                
                # Validate and correct date formats
                for date_field in ['date', 'PS_date', 'PE_date']:
                    if row_data.get(date_field) and isinstance(row_data[date_field], str):
                        row_data[date_field] = parse_date(row_data[date_field])

                # Create PolicyIssue instance
                policy_issue = PolicyIssue(**row_data)
                policy_issue.save()

            return HttpResponse('Data uploaded successfully!')

        except Exception as e:
            return HttpResponse(f'Error uploading data: {e}')

    return render(request, 'uploadxlspolicyissue.html')

def parse_date(date_str):
    formats = ['%Y-%m-%d', '%m-%d-%Y', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%Y']  # Add more formats as needed
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            pass
    #raise ValueError('Invalid date format')


def downloadData(request):
    if request.method == 'POST':
        # Get the selected table name from the HTML form
        Table = request.POST.get('Table')

        # Query the database to get the data for the selected table
        # Replace with your actual field and model name
        table_data = Users.objects.filter(table_name_field=Table)

        # Create an Excel workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Write header row
        # Replace Users with your actual model name
        headers = [field.verbose_name for field in Users._meta.fields]
        ws.append(headers)

        # Write data rows
        for row in table_data:
            ws.append([getattr(row, field.name)
                      for field in Users._meta.fields])

        # Create a response object with the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={Table}.xlsx'

        # Save the workbook to the response object
        wb.save(response)

    return render(request, 'downloaddata.html')


def export_to_excel(request):
    if request.method == 'POST':
        Table = request.POST.get('Table')

        # Check if the selected table is a valid model
        model_mapping = {
            'Users': Users,
            'CombinedInfo': CombinedInfo,
            'PolicyIssue': PolicyIssue,
            'InsuranceEnquiry': InsuranceEnquiry,
            'VehicleInformation': VehicleInformation,
            'LoanEnquiry': LoanEnquiry,
            'Document': Document,
        }

        if Table in model_mapping:
            model = model_mapping[Table]

            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{Table}_data.xlsx"'

            # Create a workbook and add a worksheet for the selected model
            workbook = Workbook()
            worksheet = workbook.active

            # Add headers to the worksheet
            headers = [field.name for field in model._meta.fields]
            worksheet.append(headers)

            # Add data to the worksheet
            queryset = model.objects.all()
            for obj in queryset:
                row_data = [str(getattr(obj, field)) for field in headers]
                worksheet.append(row_data)

            # Save the workbook to the response object
            workbook.save(response)

            return response
    

    # If there's no valid table selected or if the request is not a POST, render the downloaddata.html template
    return render(request, 'downloaddata.html')

def downloadsingletable(request):
    page = request.POST.get('param2')
    if request.method == 'GET':
        # Get the selected table name from the HTML form
        Table = request.POST.get('param1')

        # Query the database to get the data for the selected table
        # Replace with your actual field and model name
        table_data = Users.objects.filter(table_name_field=Table)

        # Create an Excel workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        # Write header row
        # Replace Users with your actual model name
        headers = [field.verbose_name for field in Users._meta.fields]
        ws.append(headers)

        # Write data rows
        for row in table_data:
            ws.append([getattr(row, field.name)
                      for field in Users._meta.fields])

        # Create a response object with the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={Table}.xlsx'

        # Save the workbook to the response object
        wb.save(response)

    return render(request, page)

    # Handle case where parameters are not provided
    return HttpResponse("Parameters are missing.", status=400)


# def export_to_excel_zip(request):
#     response = HttpResponse(content_type='application/ms-excel')
#     response['Content-Disposition'] = 'attachment; filename="exported_data.zip"'

#     # Create a zip file to contain individual Excel files for each model
#     with zipfile.ZipFile(response, 'w') as zip_file:
#         models = [Users, CombinedInfo, PolicyIssue, InsuranceEnquiry,
#                   VehicleInformation, LoanEnquiry, Document]

#         for model in models:
#             model_name = model.__name__
#             excel_file_path = f"{model_name}_data.xlsx"

#             # Create a workbook and add a worksheet for the model
#             workbook = Workbook()
#             worksheet = workbook.active

#             # Add headers to the worksheet
#             headers = [field.name for field in model._meta.fields]
#             worksheet.append(headers)

#             # Add data to the worksheet
#             queryset = model.objects.all()
#             for obj in queryset:
#                 row_data = [str(getattr(obj, field)) for field in headers]
#                 worksheet.append(row_data)

#             # Save the workbook to a temporary file
#             workbook.save(excel_file_path)

#             # Add the Excel file to the zip archive
#             zip_file.write(excel_file_path, arcname=f"{model_name}_data.xlsx")

#     return render(request, 'downloaddata.html')
